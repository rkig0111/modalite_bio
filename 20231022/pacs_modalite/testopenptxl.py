# -*- coding: utf-8 -*-
#
# I hate the documentation of openpyxl and it took me a while to undertand their stuff. So I decided to write down this code.
# Has some wrapper functions that reads all rows from the excel sheet and also a function to read a particular row.
# Add some code to the functions if you wish to do something on fly like adding values to list and sorting them later.
#
# Date: 28/09/2015

from openpyxl import load_workbook

# Reads all rows
def read_all_rows(workSheet):
    # You need to set an offset if the first row is your column headings
    for row in workSheet.iter_rows(row_offset=1):
        print("\n")
        for cell in row:
                print(type(cell.value), "   ", cell.value)
                # if type(cell.value)=='unicode':
                #         print("unicode : ", cell.value.encode('utf-8'))
                # else:
                #         print("autre : ", str(cell.value))

# Prints the data in a particular column
def read_column(workSheet, columnNumber):
    # You need to set an offset if the first row is your column headings
    for row in workSheet.iter_rows(row_offset=1):
        print(row[columnNumber].value)
        #print(row[columnNumber].value.encode('utf8'))
        


# ------------------------------ MAIN ---------------------------------------

workBook = load_workbook(filename='PACS_modalites.xlsx', read_only=True, use_iterators=True)
workSheet = workBook['PACS_modalites']
read_all_rows(workSheet)
#read_column(workSheet, 1)