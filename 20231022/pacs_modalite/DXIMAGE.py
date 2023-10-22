# -*- coding: utf-8 -*-

from openpyxl import load_workbook
#from openpyxl.utils import get_column_letter

wb = load_workbook(filename = 'DXIMAGE_ListeModalitésDeclarées.xlsx')
# ws = wb.active
ws = wb['Feuil1']
# ws = wb.get_sheet_by_name('PACS_modalites')
print("noms et nb. de feuilles : ", wb.sheetnames)
print("feuille : ", ws)
first_row = ws.rows
print("first_row : ", first_row)
rows_iter = ws.iter_rows(min_col=1, min_row=1, max_col=2, max_row=ws.max_row)
print(dir(rows_iter))
vals = [[cell.value for cell in row] for row in rows_iter]
# print(vals)

col={'Nom':0,
    'ACTIF O/N ?':1,
    'HOPITAL':2,
    'Salle(s)':3,
    'Mod.':4,
    'AET':5,
    'IP':6,
    'PORT':7 }

"""
for a in (enumerate(first_row)):
    print(a[1].value.encode('utf-8'))      # liste des colonnes"""


print(wb)
print(ws)

# for a in (enumerate(ws)):
#   print(a[1][col['Nom']].value.encode('utf-8'))


# rows_iter = ws.iter_rows(min_col=1, min_row=1, max_col=2, max_row=ws.max_row)
# vals = [[cell.value for cell in row] for row in rows_iter]

# for row in ws.iter_rows(min_row=1, max_row=5, min_col=None, max_col=None, values_only=False):
#     vals = [[cell.value for cell in row] for row in rows_iter]
#     print vals

# d = ws.cell(row = 4, column = col['résultat']+1)
# print(d.value.encode('utf8'))

