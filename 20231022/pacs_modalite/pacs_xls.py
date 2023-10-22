# -*- coding: utf-8 -*-

from openpyxl import load_workbook
#from openpyxl.utils import get_column_letter

wb = load_workbook(filename = 'PACS_modalites.xlsx')
# ws = wb.active
ws = wb['PACS_modalites']
# ws = wb.get_sheet_by_name('PACS_modalites')
# print("noms et nb. de feuilles : ", wb.sheetnames)
# print("feuille : ", ws)
first_row = ws.rows[0]

col={'Aetitle':0,
    'IP address':1,
    'Port number':2,
    'Check IP add':3,
    'Alias':4,
    'mac adresse':5,
    'Vlan':6,
    'store':7,
    'ping':8,
    'telnet':9,
    'echo dicom':10,
    'dernier envoi':11,
    'résultat':12 }


for a in (enumerate(first_row)):
    print(a[1].value.encode('utf-8'))      # liste des colonnes

for a in (enumerate(ws)):
    print(a[1][col['IP address']].value.encode('utf-8'))   # liste de toutes les adresse IP (par exemple)


# rows_iter = ws.iter_rows(min_col=1, min_row=1, max_col=2, max_row=ws.max_row)
# vals = [[cell.value for cell in row] for row in rows_iter]

# for row in ws.iter_rows(min_row=1, max_row=5, min_col=None, max_col=None, values_only=False):
#     vals = [[cell.value for cell in row] for row in rows_iter]
#     print vals

d = ws.cell(row = 4, column = col['résultat']+1)
print(d.value.encode('utf8'))

